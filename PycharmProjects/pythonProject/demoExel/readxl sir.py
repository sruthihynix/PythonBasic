import openpyxl

# Give the location of the file
path = "Sum1.xlsx"

wb_obj = openpyxl.load_workbook(path)

sheet_obj = wb_obj.active

cell_obj = sheet_obj.cell(row=1, column=1)

print(cell_obj.value)
print(sheet_obj.max_row)
# ptint total number of column
print(sheet_obj.max_column)

max_col = sheet_obj.max_column

# Loop will print all columns name
m_row = sheet_obj.max_row

# Loop will print all values
# of first column
for i in range(1, m_row + 1):
    cell_obj = sheet_obj.cell(row=i, column=1)
    print(cell_obj.value)

for i in range(1, max_col + 1):
    cell_obj = sheet_obj.cell(row=1, column=i)
    print(cell_obj.value)
