from openpyxl import Workbook,load_workbook
from openpyxl.utils import get_column_letter


wb = Workbook()
sheet = wb.active

sheet['A1'] = "Student ID"
sheet['B1'] = "Name"
sheet['C1'] = "Class"
sheet['D1'] = "Status"

limit=int(input("enter the number of students"))
for i in range(1,limit):
  sheet['A'+str(i+1)] = (input("enter the student id:"))
  sheet['B'+str(i+1)] = str(input("enter the student name:"))
  sheet['C'+str(i+1)] = int(input("enter the class:"))
  sheet['D'+str(i+1)] = str(input("enter the status :pass/fail :"))


wb.save("student_file.xlsx")